from dataclasses import dataclass
from datetime import datetime
import csv
from io import BytesIO, StringIO
import json

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.dates import parse_optional_datetime
from app.core.errors import ValidationError
from app.core.money import parse_money_to_cents
from app.domain.enums import DealType, ProductCategory, StockStatus, UseContext
from app.infrastructure.db.models import Deal, Network, Product
from app.infrastructure.repositories.catalog_repository import CatalogRepository


@dataclass(frozen=True)
class ImportIssue:
    row_number: int
    message: str


@dataclass(frozen=True)
class ImportResult:
    rows_processed: int
    products_created: int
    products_updated: int
    deals_created: int
    deals_updated: int
    networks_created: int
    issues: tuple[ImportIssue, ...]


class CsvDealImporter:
    REQUIRED_COLUMNS = frozenset(
        {
            "source_key",
            "network_code",
            "network_name",
            "sku",
            "product_name",
            "slug",
            "brand",
            "category",
            "deal_type",
            "stock_status",
            "published",
        }
    )

    def __init__(self, session: Session) -> None:
        self._session = session
        self._catalogue = CatalogRepository(session)

    @staticmethod
    def _boolean(value: str | None) -> bool:
        return (value or "").strip().lower() in {"1", "true", "yes", "y"}

    @staticmethod
    def _optional_int(value: str | None) -> int | None:
        if value is None or not value.strip():
            return None
        result = int(value.strip())
        if result < 0:
            raise ValueError("Numeric allocations cannot be negative")
        return result

    @staticmethod
    def _required(row: dict[str, str], field: str) -> str:
        value = (row.get(field) or "").strip()
        if not value:
            raise ValueError(f"{field} is required")
        return value

    def import_text(self, text: str) -> ImportResult:
        reader = csv.DictReader(StringIO(text.lstrip("\ufeff")))
        return self._import_rows(set(reader.fieldnames or []), list(reader), "CSV")

    def import_xlsx_bytes(self, contents: bytes) -> ImportResult:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
        worksheet = workbook.active
        values = worksheet.iter_rows(values_only=True)
        raw_headers = next(values, ())
        headers = [str(value).strip() if value is not None else "" for value in raw_headers]
        rows: list[dict[str, str]] = []
        for values_row in values:
            row = {
                header: self._spreadsheet_value(value)
                for header, value in zip(headers, values_row, strict=False)
                if header
            }
            if any(value for value in row.values()):
                rows.append(row)
        return self._import_rows(set(headers), rows, "spreadsheet")

    @staticmethod
    def _spreadsheet_value(value) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, bool):
            return "true" if value else "false"
        return str(value)

    def _import_rows(
        self, headers: set[str], rows: list[dict[str, str]], source_name: str
    ) -> ImportResult:
        missing = sorted(self.REQUIRED_COLUMNS - headers)
        if missing:
            raise ValidationError(f"{source_name} is missing columns: {', '.join(missing)}")
        issues = self._validate_rows(rows)
        if issues:
            return ImportResult(0, 0, 0, 0, 0, 0, tuple(issues))
        metrics = {
            "products_created": 0,
            "products_updated": 0,
            "deals_created": 0,
            "deals_updated": 0,
            "networks_created": 0,
        }
        for row in rows:
            self._apply_row(row, metrics)
        return ImportResult(rows_processed=len(rows), issues=(), **metrics)

    def _validate_rows(self, rows: list[dict[str, str]]) -> list[ImportIssue]:
        issues: list[ImportIssue] = []
        seen_source_keys: set[str] = set()
        for index, row in enumerate(rows, start=2):
            try:
                source_key = self._required(row, "source_key")
                if source_key in seen_source_keys:
                    raise ValueError("source_key is duplicated in the file")
                seen_source_keys.add(source_key)
                self._required(row, "network_code")
                self._required(row, "network_name")
                self._required(row, "sku")
                self._required(row, "product_name")
                self._required(row, "slug")
                self._required(row, "brand")
                ProductCategory(self._required(row, "category"))
                DealType(self._required(row, "deal_type"))
                StockStatus(self._required(row, "stock_status"))
                UseContext((row.get("use_context") or UseContext.BOTH.value).strip())
                for field in ("monthly_price", "cash_price", "upfront_cost"):
                    parse_money_to_cents(row.get(field))
                for field in (
                    "contract_months",
                    "data_mb",
                    "voice_minutes",
                    "sms_count",
                    "speed_mbps",
                ):
                    self._optional_int(row.get(field))
                parse_optional_datetime(row.get("start_at"))
                parse_optional_datetime(row.get("expires_at"))
                parse_optional_datetime(row.get("verified_at"))
                if row.get("specifications_json", "").strip():
                    parsed = json.loads(row["specifications_json"])
                    if not isinstance(parsed, dict):
                        raise ValueError("specifications_json must be a JSON object")
            except (ValueError, json.JSONDecodeError) as exc:
                issues.append(ImportIssue(index, str(exc)))
        return issues

    def _apply_row(self, row: dict[str, str], metrics: dict[str, int]) -> None:
        network_code = self._required(row, "network_code").lower()
        network = self._catalogue.get_network_by_code(network_code)
        if not network:
            network = Network(
                code=network_code,
                display_name=self._required(row, "network_name"),
                accent_color=(row.get("network_color") or "#7ACC00").strip(),
            )
            self._catalogue.add(network)
            metrics["networks_created"] += 1
        else:
            network.display_name = self._required(row, "network_name")
            if row.get("network_color", "").strip():
                network.accent_color = row["network_color"].strip()

        sku = self._required(row, "sku")
        product = self._catalogue.get_product_by_sku(sku)
        product_values = {
            "slug": self._required(row, "slug"),
            "name": self._required(row, "product_name"),
            "brand": self._required(row, "brand"),
            "category": ProductCategory(self._required(row, "category")),
            "description": (row.get("description") or "").strip() or None,
            "use_context": UseContext(
                (row.get("use_context") or UseContext.BOTH.value).strip()
            ),
            "primary_image_url": (row.get("image_url") or "").strip() or None,
            "specifications": json.loads(row.get("specifications_json") or "{}"),
            "is_active": True,
        }
        if not product:
            product = Product(sku=sku, **product_values)
            self._catalogue.add(product)
            metrics["products_created"] += 1
        else:
            for field, value in product_values.items():
                setattr(product, field, value)
            metrics["products_updated"] += 1

        source_key = self._required(row, "source_key")
        deal = self._session.scalar(select(Deal).where(Deal.source_key == source_key))
        deal_values = {
            "product_id": product.id,
            "network_id": network.id,
            "deal_type": DealType(self._required(row, "deal_type")),
            "monthly_price_cents": parse_money_to_cents(row.get("monthly_price")),
            "cash_price_cents": parse_money_to_cents(row.get("cash_price")),
            "upfront_cost_cents": parse_money_to_cents(row.get("upfront_cost")),
            "contract_months": self._optional_int(row.get("contract_months")),
            "data_mb": self._optional_int(row.get("data_mb")),
            "voice_minutes": self._optional_int(row.get("voice_minutes")),
            "sms_count": self._optional_int(row.get("sms_count")),
            "speed_mbps": self._optional_int(row.get("speed_mbps")),
            "starts_at": parse_optional_datetime(row.get("start_at")),
            "expires_at": parse_optional_datetime(row.get("expires_at")),
            "stock_status": StockStatus(self._required(row, "stock_status")),
            "source_document": (row.get("source_document") or "").strip() or None,
            "administrator_notes": (row.get("administrator_notes") or "").strip() or None,
            "verified_at": parse_optional_datetime(row.get("verified_at")),
            "terms_url": (row.get("terms_url") or "").strip() or None,
            "published": self._boolean(row.get("published")),
            "featured": self._boolean(row.get("featured")),
        }
        if not deal:
            deal = Deal(source_key=source_key, **deal_values)
            self._catalogue.add(deal)
            metrics["deals_created"] += 1
        else:
            for field, value in deal_values.items():
                setattr(deal, field, value)
            metrics["deals_updated"] += 1
