from __future__ import annotations

from argparse import ArgumentParser
from collections import OrderedDict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
import csv
import json
import re

from openpyxl import load_workbook

try:
    from scripts.mtn_plan_catalogue import plan_benefits
except ModuleNotFoundError:
    from mtn_plan_catalogue import plan_benefits


REQUIRED_TEMPLATE_COLUMNS = [
    "source_key",
    "network_code",
    "network_name",
    "network_color",
    "sku",
    "product_name",
    "slug",
    "brand",
    "category",
    "deal_type",
    "monthly_price",
    "cash_price",
    "upfront_cost",
    "contract_months",
    "data_mb",
    "voice_minutes",
    "sms_count",
    "speed_mbps",
    "start_at",
    "expires_at",
    "stock_status",
    "source_document",
    "administrator_notes",
    "verified_at",
    "terms_url",
    "published",
    "featured",
    "image_url",
    "description",
    "use_context",
    "specifications_json",
]


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text.strip("'").strip()


def _infer_use_context(price_plan: str) -> str:
    """Keep consumer plans personal and only explicit business plans in Business."""
    normalized = _clean(price_plan).lower()
    business_markers = ("business", "made for executive", "made to share")
    return "business" if any(marker in normalized for marker in business_markers) else "personal"


def _clean_cell(value: Any) -> str:
    if isinstance(value, str):
        value = value.strip()
        if value.lower() in {"none", "nan", "na"}:
            return ""
        return value.strip("'").strip()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return str(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return _clean(value)


def _to_iso_date(value: Any) -> str:
    if value is None or _clean(value) == "":
        return ""
    if isinstance(value, (datetime,)):
        if value.tzinfo is None:
            return value.replace(tzinfo=UTC).isoformat()
        return value.astimezone(UTC).isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        serial = float(value)
        if serial > 2000:
            try:
                dt = datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(serial))
                return dt.replace(tzinfo=UTC).isoformat()
            except (OverflowError, OSError):
                return ""
    text = _clean(value).replace("/", "-")
    text = text.replace(" ", "T")
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.fromisoformat(text)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=UTC)
            return dt.isoformat()
        except ValueError:
            try:
                dt = datetime.strptime(text, fmt)
                return dt.replace(tzinfo=UTC).isoformat()
            except ValueError:
                pass
    return ""


def _to_int(value: Any) -> str:
    text = _clean(value)
    if not text:
        return ""
    match = re.search(r"\d+", text)
    return match.group(0) if match else ""

def _slugify(text: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", text.lower())
    return re.sub(r"-+", "-", slug).strip("-")[:180]


def _normalize_status(value: str) -> str:
    value = value.strip().lower()
    if value in {"eol", "out_of_stock", "out of stock", "not available"}:
        return "out_of_stock"
    if value in {"ctb", "sel", "new", "b2b", "available", "in stock"}:
        return "in_stock"
    if value in {"limited", "subject_to_confirmation", "subject to confirmation"}:
        return "limited"
    return "subject_to_confirmation"


def _infer_category(device: str) -> str:
    lowered = device.lower()
    if lowered.startswith("use your own"):
        return "mobile_plan"
    if any(token in lowered for token in ("yellocart voucher", "point of sale", "(pos)")):
        return "accessory"
    if lowered == "crosscall t4" or lowered.startswith(("ipad", "apple ipad", "tablet", "huawei matepad", "huawei mediapad", "samsung galaxy tab", "galaxy tab", "honor pad", "oppo pad", "lenovo tab", "vodacom smart tab")):
        return "tablet"
    if any(token in lowered for token in ("laptop", "notebook", "macbook", "matebook", "ideapad", "celeron", "ryzen", "hp 15 intel", "hp n4020", "lenovo gaming")):
        return "laptop"
    if lowered.startswith("huawei mesh"):
        return "accessory"
    if any(token in lowered for token in ("router", "cpe", "mifi", "wi-fi", "wifi", "uncapped data", "fibre", "fiber", "usb adapter", "huawei e5576", "huawei e6878", "sharelink mc801a", "sh@relink mc801a", "zte mc888", "zte mf286", "zte mf296")):
        return "5g" if any(token in lowered for token in ("5g", "huawei e6878", "mc801a", "mc888")) else "lte"
    if "sim" in lowered and "phone" not in lowered and "smartphone" not in lowered:
        return "sim_only"
    return "smartphone"


def _infer_brand(device: str) -> str:
    lowered = device.lower()
    if lowered.startswith("galaxy"):
        return "Samsung"
    if lowered.startswith(("iphone", "ipad", "macbook")):
        return "Apple"
    for brand in ("apple", "samsung", "huawei", "xiaomi", "nokia", "honor", "motorola", "oppo", "vivo", "zte", "hisense", "mobicel", "cat", "crosscall", "nothing", "tozed", "lenovo", "dell", "hp", "acer", "asus", "alcatel", "jbl", "fitbit", "garmin"):
        if lowered.startswith(brand):
            return brand.title() if brand != "xiaomi" else "Xiaomi"
    if device.lower().startswith("use your own"):
        return "MTN"
    tokens = _clean(device).split()
    return tokens[0] if tokens else "MTN"


def _image_url(product_name: str, category: str) -> str:
    normalized = _clean(product_name).lower()
    if "uncapped data" in normalized or category == "broadband":
        return "/static/img/mtn-data-plan.svg"
    if normalized == "sim only" or category == "sim_only":
        return "/static/img/mtn-sim-plan.svg"
    return ""


def _build_row(
    *,
    deal_id: str,
    product_name: str,
    device_status: str,
    monthly_price: str,
    contract_term: str,
    start_at: str,
    end_at: str,
    price_plan: str,
    freebies_device: str,
    freebies_plan: str,
    source_path: str,
) -> dict[str, str]:
    benefits = plan_benefits(price_plan)
    clean_deal_id = _clean(deal_id)
    if not clean_deal_id:
        raise ValueError("Missing Deal ID")
    safe_name = _clean(product_name) or "MTN Contract Device"
    category = _infer_category(safe_name)
    brand = _infer_brand(safe_name)
    deal_type = "sim_only" if category in {"sim_only", "mobile_plan"} else "internet" if category in {"lte", "5g", "fibre"} else "accessory" if category == "accessory" else "device_contract"
    image_url = _image_url(safe_name, category)
    speed_match = re.search(r"(\d+)\s*mbps", safe_name, re.IGNORECASE)
    description = " ".join(
        part
        for part in (_clean(price_plan), _clean(freebies_device), _clean(freebies_plan), _clean_monthly_note(monthly_price))
        if part
    ).strip()

    return {
        "source_key": f"mtn:{clean_deal_id.lower()}",
        "network_code": "mtn",
        "network_name": "MTN",
        "network_color": "#FFD100",
        "sku": clean_deal_id,
        "product_name": safe_name,
        "slug": f"{_slugify(safe_name or clean_deal_id)}-{_slugify(clean_deal_id)}"[:180],
        "brand": brand,
        "category": category,
        "deal_type": deal_type,
        "monthly_price": _clean(monthly_price),
        "cash_price": "",
        "upfront_cost": "",
        "contract_months": _clean(contract_term),
        "data_mb": str(benefits["data_mb"] or ""),
        "voice_minutes": str(benefits["voice_minutes"] or ""),
        "sms_count": str(benefits["sms_count"] or ""),
        "speed_mbps": str(benefits["speed_mbps"] or (speed_match.group(1) if speed_match else "")),
        "start_at": _to_iso_date(start_at),
        "expires_at": _to_iso_date(end_at),
        "stock_status": _normalize_status(_clean(device_status)),
        "source_document": source_path,
        "administrator_notes": (
            f"Imported from MTN promo source {clean_deal_id}; contracts only"
            + ("" if image_url else "; exact product image required before publishing")
        ),
        "verified_at": datetime.now(UTC).replace(microsecond=0).isoformat(),
        "terms_url": "",
        "published": "true" if image_url else "false",
        "featured": "false",
        "image_url": image_url,
        "description": description,
        "use_context": _infer_use_context(price_plan),
        "specifications_json": json.dumps(
            {
                "deal_id": clean_deal_id,
                "source": _clean(source_path),
                "price_plan": _clean(price_plan),
                "plan_benefits": benefits,
                "freebies_device": _clean(freebies_device),
                "freebies_plan": _clean(freebies_plan),
            },
            separators=(",", ":"),
        ),
    }


def _clean_monthly_note(price: str) -> str:
    value = _clean(price)
    if not value:
        return ""
    return f"Monthly subscription: {value}"


def _iter_mtn_promo_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []

    for sheet_name in ("PromoList", "PromoList (2)"):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        sheet_rows = sheet.iter_rows(values_only=True)
        header_row = None
        for raw_row in sheet_rows:
            candidate = _clean(raw_row[0]) if raw_row else ""
            if candidate.lower() in {"'deal id'", "deal id", "dealid", "deal_id"}:
                header_row = [str(_clean(v)).lower() for v in raw_row]
                break
        if not header_row:
            continue
        header_index = {name.replace("'", "").strip().lower(): i for i, name in enumerate(header_row)}
        def get(row_data, label: str) -> str:
            index = header_index.get(label, -1)
            if index == -1 or index >= len(row_data):
                return ""
            return _clean_cell(row_data[index])

        for raw_row in sheet_rows:
            if not any(v is not None and str(v).strip() for v in raw_row):
                continue
            if not _clean(get(raw_row, "deal id")):
                continue
            monthly = get(raw_row, "total subscription incl vat")
            if not monthly:
                continue
            rows.append(
                _build_row(
                    deal_id=get(raw_row, "deal id"),
                    product_name=get(raw_row, "oem and device"),
                    device_status=get(raw_row, "device status"),
                    monthly_price=monthly,
                    contract_term=get(raw_row, "contract term"),
                    start_at=get(raw_row, "promo start date (mm/dd/yyyy)"),
                    end_at=get(raw_row, "promo end date (mm/dd/yyyy)"),
                    price_plan=get(raw_row, "price plan"),
                    freebies_device=get(raw_row, "freebies description 1\n(devices)"),
                    freebies_plan=get(raw_row, "freebie description 2\n(priceplan)"),
                    source_path=str(path),
                )
            )
    return rows


def _iter_cheat_sheet_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "August 2026" not in workbook.sheetnames:
        return []
    sheet = workbook["August 2026"]
    sheet_rows = sheet.iter_rows(values_only=True)
    header_row = _clean_row = [str(_clean(v)).lower() for v in next(sheet_rows)]
    idx = {name.lower(): i for i, name in enumerate(header_row)}
    rows: list[dict[str, str]] = []
    for row in sheet_rows:
        if not any(v is not None and str(v).strip() for v in row):
            continue
        deal_id = _clean_cell(row[idx.get("deal id reference", -1)])
        monthly = _clean_cell(row[idx.get("pricing subscription", -1)])
        if not deal_id or not monthly:
            continue
        rows.append(
            _build_row(
                deal_id=deal_id,
                product_name=_clean_cell(row[idx.get("device", -1)]),
                device_status="CTB",
                monthly_price=monthly,
                contract_term=_clean_cell(row[idx.get("contract term", -1)]),
                start_at=_clean_cell(row[idx.get("promo start date", -1)]),
                end_at=_clean_cell(row[idx.get("promo end date", -1)]),
                price_plan=_clean_cell(row[idx.get("price plan", -1)]),
                freebies_device=_clean_cell(row[idx.get("freebies - description", -1)]),
                freebies_plan="",
                source_path=str(path),
            )
        )
    return rows


def main() -> int:
    parser = ArgumentParser(
        description="Convert MTN promo inventories to the app's deal-import CSV format."
    )
    parser.add_argument("paths", nargs="+", help="Path to MTN Excel files (.xlsx)")
    parser.add_argument(
        "--output",
        default="data/mtn-contracts-import.csv",
        help="Output CSV path (required import template format)",
    )
    args = parser.parse_args()

    collected: list[dict[str, str]] = []
    for raw_path in args.paths:
        path = Path(raw_path)
        if path.name.lower().startswith("outright pricing"):
            continue
        if path.suffix.lower() != ".xlsx":
            continue
        if "cheat sheet" in path.name.lower():
            collected.extend(_iter_cheat_sheet_rows(path))
        else:
            collected.extend(_iter_mtn_promo_rows(path))

    if not collected:
        raise SystemExit("No contract rows found in the supplied files.")

    # Keep the latest duplicate source keys only, in case files overlap.
    ordered = OrderedDict()
    for row in collected:
        ordered[row["source_key"]] = row

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=REQUIRED_TEMPLATE_COLUMNS)
        writer.writeheader()
        writer.writerows(ordered.values())
    print(f"Exported {len(ordered)} contract rows to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
